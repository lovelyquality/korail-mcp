import csv
import json
import os
from pathlib import Path
from typing import Optional

import httpx
import openpyxl
from dotenv import load_dotenv
from mcp.server.fastmcp import FastMCP

load_dotenv(encoding="utf-8-sig")
API_KEY = os.getenv("DATA_GO_KR_API_KEY")
ODCLOUD_BASE = "https://api.odcloud.kr/api"
DATA_DIR = Path(__file__).parent / "data"

mcp = FastMCP("korail-network")
_cache: dict = {}

# ─── odcloud helpers ──────────────────────────────────────────────────────────

def _odcloud_get(path: str, page: int = 1, per_page: int = 1000) -> dict:
    r = httpx.get(
        f"{ODCLOUD_BASE}{path}",
        params={"serviceKey": API_KEY, "page": page, "perPage": per_page},
        timeout=20,
    )
    return r.json()


def _odcloud_load_all(path: str) -> list:
    all_data, page = [], 1
    while True:
        body = _odcloud_get(path, page=page, per_page=1000)
        data = body.get("data", [])
        all_data.extend(data)
        if len(all_data) >= body.get("totalCount", 0) or not data:
            break
        page += 1
    return all_data


# ─── local file loaders ───────────────────────────────────────────────────────

def _load_csv(filename: str) -> list[dict]:
    with open(DATA_DIR / filename, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _load_station_distance_index() -> dict:
    """역간최단거리 CSV → {출발역명: [행 목록]} 인덱스"""
    data = _load_csv("station_distance.csv")
    idx: dict[str, list] = {}
    for row in data:
        dep = row.get("출발역명", "")
        if dep not in idx:
            idx[dep] = []
        idx[dep].append(row)
    return idx


def _parse_line_distance_sheet(ws) -> tuple[list, dict]:
    """삼각행렬 형태의 운행거리 시트를 파싱 → (역명목록, {(출발,도착): km 딕셔너리})

    형식: 행 k의 col k에 역명, col k+1 이후에 다음 역까지의 누적거리.
    """
    rows = list(ws.iter_rows(values_only=True))

    # 대각선 시작 행: col=0에 문자열이 있는 첫 행
    start_row = None
    for row_idx, row in enumerate(rows):
        if row and len(row) > 0 and isinstance(row[0], str) and row[0].strip():
            start_row = row_idx
            break

    if start_row is None:
        return [], {}

    # 역명 추출 (대각선: rows[start_row+k][k] = 역명)
    station_names: list[str] = []
    for k in range(len(rows) - start_row):
        ridx = start_row + k
        if ridx >= len(rows):
            break
        row = rows[ridx]
        if k >= len(row) or row[k] is None:
            break
        val = row[k]
        if isinstance(val, str) and val.strip():
            station_names.append(val.strip())
        else:
            break

    # 거리 딕셔너리 구축 (양방향)
    distances: dict[tuple, float] = {}
    for i, from_st in enumerate(station_names):
        row = rows[start_row + i]
        for j in range(i + 1, len(station_names)):
            to_st = station_names[j]
            dist = row[j] if j < len(row) else None
            if dist is not None and isinstance(dist, (int, float)) and float(dist) > 0:
                d = round(float(dist), 1)
                distances[(from_st, to_st)] = d
                distances[(to_st, from_st)] = d

    return station_names, distances


def _load_operation_distance_all() -> dict:
    """철도운행거리_전체 XLSX 전체 시트 파싱 → {시트명: {stations, distances}}"""
    wb = openpyxl.load_workbook(
        DATA_DIR / "operation_distance_all.xlsx", read_only=True, data_only=True
    )
    result = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == "000000":
            continue
        ws = wb[sheet_name]
        stations, distances = _parse_line_distance_sheet(ws)
        if stations:
            result[sheet_name] = {"stations": stations, "distances": distances}
    wb.close()
    return result


# ─── cache helper ─────────────────────────────────────────────────────────────

def _get(key: str, loader):
    if key not in _cache:
        _cache[key] = loader()
    return _cache[key]


# ─── 메타데이터 헬퍼 ──────────────────────────────────────────────────────────

def _make_meta(dataset: str, ref_date: str) -> dict:
    """출처·기준일 메타 딕셔너리 생성."""
    return {
        "출처": "한국철도공사 공공데이터포털 (data.go.kr)",
        "데이터셋": dataset,
        "데이터기준일": ref_date,
    }


def _wrap_list(data: list, dataset: str, ref_date: str) -> dict:
    """리스트 결과를 메타와 함께 딕셔너리로 반환."""
    return {
        "data": data,
        "건수": len(data),
        "_meta": _make_meta(dataset, ref_date),
    }


# ─── Tool 1: 노선 검색 ────────────────────────────────────────────────────────

ROUTE_PATH = "/15138455/v1/uddi:8aea9f31-7bd0-4870-9553-8f0fb49075ec"


@mcp.tool()
def search_routes(
    query: Optional[str] = None,
    electric_only: Optional[bool] = None,
) -> list[dict]:
    """전국 철도 노선 정보를 검색합니다. (총 2,146개)

    ⚠️ 여기서 "노선"은 물리적 선로(경부선·호남선 등)가 아니라 열차 운행계통 코드입니다.
    경부선 선로 하나에도 다양한 경유 패턴의 운행계통이 수백 개 존재합니다.
    "비전기 노선이 몇 개냐" 같은 인프라 기준 질문에는 수치가 과대 계산됩니다.

    Args:
        query: 노선코드(ROUT_CD) 또는 노선명(ROUT_NM) 검색어 (부분 일치). 없으면 전체 반환.
               ※ 노선명에 "KTX" 문자열 없음. 고속선은 "고속", "경부고속" 등으로 표기됨.
        electric_only: True면 전기동력차 운행 노선만, False면 비전기 노선만. None이면 전체.

    Returns:
        노선코드(ROUT_CD), 노선명(ROUT_NM), 전기동력차운행여부(ELC_LCM_RUN_FLG) 목록.
    """
    data = _get("routes", lambda: _odcloud_load_all(ROUTE_PATH))
    result = data
    if query:
        q = query.upper()
        result = [
            r for r in result
            if q in str(r.get("노선코드(ROUT_CD)", "")).upper()
            or query in str(r.get("노선명(ROUT_NM)", ""))
        ]
    if electric_only is not None:
        flag = "Y" if electric_only else "N"
        result = [r for r in result if r.get("전기동력차운행여부(ELC_LCM_RUN_FLG)") == flag]
    return _wrap_list(result, "한국철도공사_철도운영정보_노선정보", "2024.09.01")


# ─── Tool 2: 역간 최단거리 ────────────────────────────────────────────────────

@mcp.tool()
def get_station_distance(
    from_station: str,
    to_station: Optional[str] = None,
    current_only: bool = True,
) -> list[dict]:
    """두 역 간 최단 운행거리(km)를 조회합니다. 여객·화물 거리를 구분하며 경유역 정보 포함.

    Args:
        from_station: 출발역명 (정확한 역명 권장, 부분 일치도 지원. 예: "서울", "부산").
        to_station: 도착역명 (부분 일치). 없으면 출발역에서 출발하는 모든 구간 거리 반환.
        current_only: True면 현재 유효 데이터(적용종료일자=9999-12-31)만 반환. 기본 True.

    Returns:
        출발역명, 도착역명, 여객최단운행거리(km), 화물운행거리(km), 구간거리내용(경유역), 적용기간.
        ※ 총 220,782건. 최초 호출 시 로딩에 수 초가 소요됩니다. 최대 200건 반환.
        ※ 출발역명 기준 인덱스만 존재. "특정 역에 도착하는 모든 경로" 역방향 조회는 미지원.
        ※ 여객최단운행거리는 운임 계산 기준 거리로, 실제 열차 운행거리(XLSX 기준)와 다를 수 있음.
           예: 서울↔부산 여객최단 441.7 km vs KTX 운행거리 417.4 km (고속선 vs 운임 기준선 차이).
    """
    idx = _get("station_distance_idx", _load_station_distance_index)

    # 정확 일치 우선, 없으면 부분 일치
    rows = idx.get(from_station, [])
    if not rows:
        for k, v in idx.items():
            if from_station in k:
                rows = rows + v

    if to_station:
        rows = [r for r in rows if to_station in r.get("도착역명", "")]
    if current_only:
        rows = [r for r in rows if str(r.get("적용종료일자", "")).strip() == "9999-12-31"]

    return _wrap_list(rows[:200], "한국철도공사_철도운영정보_역간최단거리", "2024.12.31")


# ─── Tool 3: 화물 최저운임 ────────────────────────────────────────────────────

FARE_PATH = "/15153539/v1/uddi:69cf6c1d-fbff-4981-a65d-b9e197e14911"


@mcp.tool()
def get_freight_minimum_fare(
    fare_type: Optional[str] = None,
    classification_no: Optional[str] = None,
    current_only: bool = False,
) -> list[dict]:
    """화물 운송 최저운임 기준 정보를 조회합니다. (총 7건 — 역사적 이력 포함)

    Args:
        fare_type: 운임요금유형 (예: "최저운임"). 없으면 전체.
        classification_no: 분류번호 (예: "10", "20", "30"). 없으면 전체.
        current_only: True면 적용종료일자가 2100년 이후인 현재 유효 운임만 반환.

    Returns:
        운임요금유형, 분류번호, 분류번호내용, 적용최저운임(원), 적용시작일자, 적용종료일자.
    """
    data = _get("freight_fare", lambda: _odcloud_load_all(FARE_PATH))
    result = data
    if fare_type:
        result = [r for r in result if fare_type in str(r.get("운임요금유형", ""))]
    if classification_no:
        result = [r for r in result if str(r.get("분류번호", "")).strip() == classification_no.strip()]
    if current_only:
        result = [r for r in result if str(r.get("적용종료일자", "")) >= "2100"]
    return _wrap_list(result, "한국철도공사_철도운영정보_운임", "2025.11.07")


# ─── Tool 4: 임율 조회 ────────────────────────────────────────────────────────

RATE_PATH = "/15153571/v1/uddi:8b1350c1-711c-422a-b68d-e4e27ed31509"


@mcp.tool()
def get_freight_rate(
    category: Optional[str] = None,
    classification_no: Optional[str] = None,
    current_only: bool = True,
) -> list[dict]:
    """철도 화물 임율(ton-km 기준 운임 요율) 정보를 조회합니다. (전체 249건, 현재 유효 약 123건)

    Args:
        category: 신청구분 (예: "일반", "컨테이너"). 없으면 전체.
        classification_no: 분류번호 (예: "10", "1001", "1021"). 없으면 전체.
        current_only: True면 현재 적용 중인 임율만 반환 (적용종료일자 2100 이후). 기본 True.
                      ※ 분류번호 1xxx는 컨테이너 임율. 10~80번대는 일반 화물 임율.

    Returns:
        신청구분, 분류번호, 분류번호내용(화물 유형 설명), 적용임율(원/ton-km), 컨테이너규격내용, 적용기간.
    """
    data = _get("freight_rate", lambda: _odcloud_load_all(RATE_PATH))
    result = data
    if category:
        result = [r for r in result if category in str(r.get("신청구분", ""))]
    if classification_no:
        result = [r for r in result if str(r.get("분류번호", "")).strip() == classification_no.strip()]
    if current_only:
        result = [r for r in result if str(r.get("적용종료일자", "")) >= "2100"]
    return _wrap_list(result, "한국철도공사_철도운영정보_임율", "2024.07.01")


# ─── Tool 5: 전동차 세그먼트 정보 ─────────────────────────────────────────────

@mcp.tool()
def get_segment_info(
    segment_code: Optional[str] = None,
    region: Optional[str] = None,
    station: Optional[str] = None,
) -> dict:
    """철도 전동차 세그먼트(구간) 정보를 조회합니다.
    세그먼트는 노선을 운행 분석 최소 단위로 분리한 구간입니다.

    Args:
        segment_code: 세그먼트코드 (예: "100", "200", "300").
        region: 운행지역본부 부분 일치 (예: "서울", "대전", "수도권서부", "부산").
        station: 세그먼트 시작역·종료역·경유역명 부분 일치 (예: "서울", "대전").

    Returns:
        basic: 세그먼트 기본 정보 (코드, 기간, 지역본부, 시작역, 종료역, 거리(km)).
        detail: 해당 세그먼트 구성 역 상세 (역명, 역일련번호, 누적거리).
        total_segments: 매칭된 세그먼트 수.
    """
    basics = _get("segment_basic", lambda: _load_csv("segment_basic.csv"))
    details = _get("segment_detail", lambda: _load_csv("segment_detail.csv"))

    filtered = basics
    if segment_code:
        filtered = [b for b in filtered if str(b.get("세그먼트코드", "")).strip() == segment_code.strip()]
    if region:
        filtered = [b for b in filtered if region in str(b.get("운행지역본부구분", ""))]
    if station:
        filtered = [
            b for b in filtered
            if station in str(b.get("세그먼트시작역명", ""))
            or station in str(b.get("세그먼트종료역명", ""))
        ]

    # 경유역에서도 검색 (station 지정 시 detail에서 추가 탐색)
    if station and not segment_code:
        extra_codes = {
            str(d.get("세그먼트코드", "")).strip()
            for d in details
            if station in str(d.get("역명", ""))
        }
        existing_codes = {str(b.get("세그먼트코드", "")).strip() for b in filtered}
        extra = [b for b in basics if str(b.get("세그먼트코드", "")).strip() in extra_codes - existing_codes]
        filtered = filtered + extra

    result_codes = {str(b.get("세그먼트코드", "")).strip() for b in filtered}
    filtered_details = [d for d in details if str(d.get("세그먼트코드", "")).strip() in result_codes]

    return {
        "basic": filtered[:100],
        "detail": filtered_details[:500],
        "total_segments": len(filtered),
        "_meta": _make_meta("한국철도공사_철도운영정보_전동차구간 정보", "2025.11.07"),
    }


# ─── Tool 6: 노선별 운행거리 ──────────────────────────────────────────────────

@mcp.tool()
def get_operation_distance(
    line_name: Optional[str] = None,
    from_station: Optional[str] = None,
    to_station: Optional[str] = None,
) -> dict:
    """전국 철도 노선별 역간 운행거리를 조회합니다. (20개 노선 그룹)

    열차가 실제로 주행하는 선로 거리 기준입니다. 운임 계산용 여객최단운행거리와 다를 수 있습니다.
    (예: 서울↔부산 KTX 운행거리 417.4 km, 여객최단운행거리 441.7 km)

    ⚠️ 이 데이터는 철도운행거리_전체 XLSX(노선별 삼각행렬)에 수록된 역만 포함합니다.
       KTX 전용선 경유 역(경부KTX: 서울·영등포·광명·천안아산·오송·대전·김천구미·동대구·부산 등)만 있고,
       같은 KTX가 경유해도 행신·수원처럼 별도 인입선·재래선 역은 미포함일 수 있습니다.

    Args:
        line_name: 노선명 부분 일치 (예: "경부", "호남", "전라", "강릉", "영동", "중앙", "태백").
                   없으면 전체 노선 그룹 목록 반환.
        from_station: 출발역명 (정확 일치). 해당 역에서 출발하는 모든 거리 반환.
        to_station: 도착역명 (정확 일치). from_station과 함께 지정 시 두 역 간 거리 반환.

    Returns:
        line_name 없음: available_lines 목록.
        line_name만: 매칭 노선과 역 목록.
        from_station 추가: 출발역에서 각 역까지의 거리 목록(거리 오름차순 정렬).
        from+to 모두: 두 역 간 거리(km).
        ※ 최초 호출 시 XLSX 파싱에 수 초 소요됩니다.
    """
    all_lines = _get("operation_distance_all", _load_operation_distance_all)

    _meta = _make_meta("한국철도공사_철도운행거리_전체", "2024.09.01")

    if not line_name:
        return {
            "available_lines": list(all_lines.keys()),
            "total_lines": len(all_lines),
            "usage": "line_name 파라미터로 노선명 지정 시 상세 거리 조회 가능",
            "_meta": _meta,
        }

    matched = {name: info for name, info in all_lines.items() if line_name in name}

    if not matched:
        return {
            "error": f"'{line_name}' 노선을 찾을 수 없습니다.",
            "available_lines": list(all_lines.keys()),
            "_meta": _meta,
        }

    if from_station and to_station:
        results = []
        for name, info in matched.items():
            dist = info["distances"].get((from_station, to_station))
            if dist is not None:
                results.append({
                    "노선그룹": name,
                    "출발역": from_station,
                    "도착역": to_station,
                    "운행거리(km)": dist,
                })
        if not results:
            all_stations = list({st for info in matched.values() for st in info["stations"]})
            return {
                "error": f"'{from_station}' → '{to_station}' 거리 없음 (역명 확인 필요)",
                "available_stations": all_stations,
                "_meta": _meta,
            }
        return {"distance_result": results, "_meta": _meta}

    if from_station:
        results = []
        for name, info in matched.items():
            for (f, t), dist in info["distances"].items():
                if f == from_station:
                    results.append({"노선그룹": name, "도착역": t, "운행거리(km)": dist})
        if not results:
            all_stations = list({st for info in matched.values() for st in info["stations"]})
            return {
                "error": f"'{from_station}' 역을 찾을 수 없습니다.",
                "available_stations": all_stations,
                "_meta": _meta,
            }
        return {
            "from_station_distances": sorted(results, key=lambda x: x["운행거리(km)"]),
            "_meta": _meta,
        }

    return {
        "matched_lines": list(matched.keys()),
        "stations": {name: info["stations"] for name, info in matched.items()},
        "_meta": _meta,
    }


# ─── Tool 7: KTX 노선별 역정보 ────────────────────────────────────────────────

KTX_STATION_PATH = "/15127571/v1/uddi:ab540482-aa65-411d-908b-c961aadae08b"


@mcp.tool()
def get_ktx_stations(
    line_name: Optional[str] = None,
    station_name: Optional[str] = None,
) -> list[dict]:
    """KTX 노선별 역 정보(역명, 도로명주소, 정차 순번)를 조회합니다. (총 102건)

    Args:
        line_name: KTX 노선명 부분 일치 (예: "경부선", "호남선", "경전선"). 없으면 전체.
        station_name: 역명 부분 일치 (예: "서울", "부산", "광명"). 없으면 필터 없음.

    Returns:
        고속철도명, 철도운영기관, 노선명, 순번, 역명, 주소(도로명) 목록.
    """
    data = _get("ktx_stations", lambda: _odcloud_load_all(KTX_STATION_PATH))
    result = data
    if line_name:
        result = [r for r in result if line_name in str(r.get("노선명", ""))]
    if station_name:
        result = [r for r in result if station_name in str(r.get("역명", ""))]
    return _wrap_list(result, "한국철도공사_KTX 노선별 역정보", "2025.11.21")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--transport", default="stdio", choices=["stdio", "sse"])
    parser.add_argument("--port", type=int, default=8007)
    args = parser.parse_args()
    if args.transport == "sse":
        mcp.settings.host = "0.0.0.0"
        mcp.settings.port = args.port
        mcp.settings.transport_security = None
    mcp.run(transport=args.transport)
